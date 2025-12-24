import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def create_users_excel(users: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Create Excel file with users data.
    Columns: №, FIO, Telefon, Username, Holat
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["№", "FIO", "Telefon", "Username", "Holat", "Ro'yxatdan o'tgan"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=user.get('full_name', '')).border = thin_border
        ws.cell(row=row, column=3, value=user.get('phone', '')).border = thin_border
        ws.cell(row=row, column=4, value=user.get('username', '')).border = thin_border

        status = "Bloklangan" if user.get('is_blocked') else "Faol"
        ws.cell(row=row, column=5, value=status).border = thin_border

        created_at = user.get('created_at')
        if isinstance(created_at, datetime):
            created_at = created_at.strftime("%Y-%m-%d %H:%M")
        ws.cell(row=row, column=6, value=created_at).border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_poll_results_excel(poll_name: str, voters: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Create Excel file with poll voting results.
    Columns: №, FIO, Username, Telefon, Ro'yxatdan o'tgan, Ovoz bergan vaqt, Tanlagan variant
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ovozlar"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value=f"So'rovnoma: {poll_name}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Report date
    ws.merge_cells('A2:G2')
    ws.cell(row=2, column=1, value=f"Hisobot sanasi: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    ws['A2'].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["№", "F.I.O", "Username", "Telefon", "Ro'yxatdan o'tgan", "Ovoz bergan vaqt", "Tanlagan variant"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    for row, voter in enumerate(voters, 5):
        ws.cell(row=row, column=1, value=row-4).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=2, value=voter.get('full_name', '-')).border = thin_border

        username = voter.get('username')
        ws.cell(row=row, column=3, value=f"@{username}" if username else "-").border = thin_border

        ws.cell(row=row, column=4, value=voter.get('phone') or '-').border = thin_border

        registered_at = voter.get('registered_at')
        if isinstance(registered_at, datetime):
            registered_at = registered_at.strftime("%d.%m.%Y %H:%M")
        ws.cell(row=row, column=5, value=registered_at or '-').border = thin_border

        voted_at = voter.get('voted_at')
        if isinstance(voted_at, datetime):
            voted_at = voted_at.strftime("%d.%m.%Y %H:%M")
        ws.cell(row=row, column=6, value=voted_at or '-').border = thin_border

        ws.cell(row=row, column=7, value=voter.get('option_name', '-')).border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 25

    # Summary row
    summary_row = len(voters) + 6
    ws.merge_cells(f'A{summary_row}:G{summary_row}')
    ws.cell(row=summary_row, column=1, value=f"Jami ovozlar: {len(voters)}")
    ws[f'A{summary_row}'].font = Font(bold=True)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_statistics_excel(stats: Dict[str, Any]) -> io.BytesIO:
    """Create Excel file with bot statistics."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistika"

    # Style
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:B1')
    title_cell = ws.cell(row=1, column=1, value="Bot Statistikasi")
    title_cell.font = Font(bold=True, size=14)

    date_cell = ws.cell(row=2, column=1, value=f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # Users section
    ws.cell(row=4, column=1, value="Foydalanuvchilar").font = header_font
    ws.cell(row=5, column=1, value="Jami:").border = thin_border
    ws.cell(row=5, column=2, value=stats.get('total_users', 0)).border = thin_border
    ws.cell(row=6, column=1, value="Bugun:").border = thin_border
    ws.cell(row=6, column=2, value=stats.get('today_users', 0)).border = thin_border
    ws.cell(row=7, column=1, value="Bu hafta:").border = thin_border
    ws.cell(row=7, column=2, value=stats.get('week_users', 0)).border = thin_border
    ws.cell(row=8, column=1, value="Bu oy:").border = thin_border
    ws.cell(row=8, column=2, value=stats.get('month_users', 0)).border = thin_border

    # Polls section
    ws.cell(row=10, column=1, value="Sorovnomalar").font = header_font
    ws.cell(row=11, column=1, value="Jami:").border = thin_border
    ws.cell(row=11, column=2, value=stats.get('total_polls', 0)).border = thin_border
    ws.cell(row=12, column=1, value="Faol:").border = thin_border
    ws.cell(row=12, column=2, value=stats.get('active_polls', 0)).border = thin_border

    # Votes section
    ws.cell(row=14, column=1, value="Ovozlar").font = header_font
    ws.cell(row=15, column=1, value="Jami:").border = thin_border
    ws.cell(row=15, column=2, value=stats.get('total_votes', 0)).border = thin_border
    ws.cell(row=16, column=1, value="Bugun:").border = thin_border
    ws.cell(row=16, column=2, value=stats.get('today_votes', 0)).border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
